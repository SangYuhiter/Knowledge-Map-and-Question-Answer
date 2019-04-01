# -*- coding: utf-8 -*-
"""
@File  : QuestionTemplate.py
@Author: SangYu
@Date  : 2019/3/16 10:27
@Desc  : 问题模板
"""
from Log.Logger import MyLog
import sys
from openpyxl import load_workbook
import re
import pickle


# 求列表元素的子集(二进制法)
# print(get_subset_binary([1, 2, 3]))
def get_subset_binary(item: list) -> list:
    """
    求列表元素的子集(二进制法)
    :param item: 元素列表
    :return: 子集的列表
    """
    n = len(item)
    sub_set = []
    for i in range(2 ** n):
        tem_set = []
        for j in range(n):
            if (i >> j) % 2 == 1:
                tem_set.append(item[j])
        sub_set.append(tem_set)
    return sub_set


# 通过excel表格加载表格内容
# noinspection PyProtectedMember
def load_table_content(file_path: str):
    """
    通过excel表格加载表格内容
    :param file_path:
    :return:
    """
    function_logger = MyLog(logger=sys._getframe().f_code.co_name).getlog()
    # 加载excel表格
    function_logger.info("加载表格:%s" % file_path.split("\\")[-1])
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    sheet_first = wb.get_sheet_by_name(sheet_names[0])
    table_head = []
    for item in range(1, sheet_first.max_column + 1):
        table_head.append(sheet_first.cell(row=1, column=item).value)
    function_logger.debug("表头:%s" % str(table_head))
    table_attr = {}
    for i_column in range(1, sheet_first.max_column + 1):
        column_name = sheet_first.cell(row=1, column=i_column).value
        column_value = set()
        for i_row in range(2, sheet_first.max_row + 1):
            column_value.add(sheet_first.cell(row=i_row, column=i_column).value)
        table_attr[column_name] = str(list(column_value))
    for key in table_attr:
        function_logger.debug(key)
        value_list = [value.replace("'", "").strip() for value in table_attr[key][1:-1].split(",")]
        value_list.sort()
        function_logger.debug("列表长度:%d" % len(value_list))
        function_logger.debug(str(value_list))
    function_logger.info("加载表格:%s完成!" % file_path.split("\\")[-1])


# 通过规定字段构造模板
def build_template_by_fields(template_path: str):
    """
    通过规定字段构造模板
    :param template_path: 模板路径
    :return:
    """
    # 根据表名分别进行模板构造
    if "admission_plan" in template_path:
        build_template_by_fields_plan(template_path)
    elif "admission_score_pro" in template_path:
        build_template_by_fields_score_pro(template_path)
    elif "admission_score_major" in template_path:
        build_template_by_fields_score_major(template_path)


# 构造招生计划问题模板
def build_template_by_fields_plan(template_path: str):
    """
    构造招生计划问题模板
    :param template_path: 模板路径
    :return:
    """
    fields_question_condition = ["school 学校", "year 年份", "major 专业", "district 省份", "classy 类别"]
    fields_question_target = ["numbers 招生人数 招生计划 招多少人 招生计划是多少 招生人数是多少"]
    template_sentence_questions = ["(school)(year)(major)(district)(classy)(numbers)"]
    template_sentence_answers = ["(school)(year)(major)(district)招收(classy)(numbers)人"]
    build_template_by_infos(template_path, fields_question_condition, fields_question_target,
                            template_sentence_questions, template_sentence_answers)


# 构造录取分数分省问题模板
def build_template_by_fields_score_pro(template_path: str):
    """
    构造录取分数分省问题模板
    :param template_path: 模板路径
    :return:
    """
    fields_question_condition = ["school 学校", "year 年份", "district 省份 地区", "batch 批次", "classy 类别"]
    fields_question_target = ["line 分数线 录取分数 多少分 录取分数是多少 分数线是多少"]
    template_sentence_questions = ["(school)(year)(district)(batch)(classy)(line)"]
    template_sentence_answers = ["(school)(year)(district)(batch)(classy)录取分数是(line)"]
    build_template_by_infos(template_path, fields_question_condition, fields_question_target,
                            template_sentence_questions, template_sentence_answers)


# 构造录取分数分专业问题模板
def build_template_by_fields_score_major(template_path: str):
    """
    构造录取分数分专业问题模板
    :param template_path: 模板路径
    :return:
    """
    fields_question_condition = ["school 学校", "year 年份", "major 专业", "district 省份 地区", "classy 类别"]
    fields_question_target = ["highest 最高分 最高分是多少",
                              "average 平均分 平均分是多少",
                              "lowest 最低分 最低分是多少 分数线 分数线是多少",
                              "amount 录取人数 录取多少人"]
    template_sentence_questions = ["(school)(year)(major)(district)(classy)(highest)",
                                   "(school)(year)(major)(district)(classy)(average)",
                                   "(school)(year)(major)(district)(classy)(lowest)",
                                   "(school)(year)(major)(district)(classy)(amount)"]
    template_sentence_answers = ["(school)(year)(major)(district)(classy)最高分是(highest)",
                                 "(school)(year)(major)(district)(classy)平均分是(average)",
                                 "(school)(year)(major)(district)(classy)最低分是(lowest)",
                                 "(school)(year)(major)(district)(classy)录取人数是(amount)"]
    build_template_by_infos(template_path, fields_question_condition, fields_question_target,
                            template_sentence_questions, template_sentence_answers)


# 通过提供的信息构造问题模板
# noinspection PyProtectedMember,PyShadowingNames
def build_template_by_infos(template_path: str, fields_question_condition: list, fields_question_target: list,
                            template_sentence_questions: list, template_sentence_answers: list):
    """
    通过提供的信息构造问题模板
    :param template_path: 模板路径
    :param fields_question_condition: 问句条件词，例["school 学校", "year 年份", "major 专业", "district 省份", "classy 类别"]
    :param fields_question_target: 问句目标词，例["numbers 招生人数 招生计划 招多少人 招生计划是多少 招生人数是多少"]
    :param template_sentence_questions: 模板问题句
    :param template_sentence_answers: 模板答案句
    :return:
    """
    function_logger = MyLog(logger=sys._getframe().f_code.co_name).getlog()
    function_logger.info("开始构造%s的问题模板..." % template_path.split("\\")[-1])
    # 使用pickle存储模板文件
    template_dict = {}
    template_dict["fq_conditon"] = fields_question_condition
    template_dict["fq_target"] = fields_question_target
    template_dict["ts_answers"] = template_sentence_answers
    build_question_sentences = []
    # 获取问句条件词的英文字段
    template_fields_en = []
    for fq_condition in fields_question_condition:
        template_fields_en.append(fq_condition.split(" ")[0])
    # 利用问句条件词英文字段构造子集，使用替换的方法构造所有全排列的问句
    fields_en_subset = get_subset_binary(template_fields_en)
    print(len(fields_en_subset))
    for i_question in range(len(template_sentence_questions)):
        # 查询当前模板问句包含多少问句目标词
        match_question_target = []
        for fq_target in fields_question_target:
            if fq_target.split(" ")[0] in template_sentence_questions[i_question]:
                match_question_target = fq_target.split(" ")
                # 找到一个问句目标词即退出，默认问句中只有一个问句目标词
                break
        # 对问句目标词中对应的每一个询问方式进行替换
        target_word = match_question_target[0]
        for question_mode in match_question_target[1:]:
            # 对子集中的每一个集合进行替换
            for subset in fields_en_subset:
                sentence = template_sentence_questions[i_question].replace("(" + target_word + ")", question_mode) \
                           + "--" + str(i_question)
                # 子集为空，不缺省参数
                if not subset:
                    build_question_sentences.append(sentence)
                # 子集为原集合，不添加
                elif len(subset) == len(template_fields_en):
                    continue
                # 子集有缺省，去除子集中的元素后再添加
                else:
                    for field_en in subset:
                        sentence = sentence.replace("(" + field_en + ")", "")
                    build_question_sentences.append(sentence)
    template_dict["ts_questions"] = build_question_sentences
    with open(template_path, "wb")as p_file:
        pickle.dump(template_dict, p_file)
    function_logger.info("%s的问题模板构建完成!" % template_path.split("\\")[-1])


# 通过模板类型（槽位）构造MySQL语句
# noinspection PyProtectedMember
def build_mysql_string_by_template(template_question: str, template_question_type: str) -> str:
    """
    通过模板类型（槽位）构造MySQL语句
    :param template_question: 模板问句
    :param template_question_type: 模板问句类型
    :return: 对应的mysql语句
    """
    function_logger = MyLog(logger=sys._getframe().f_code.co_name).getlog()
    function_logger.info("开始构造MySQL语句...")
    search_table = template_question_type
    # 提取模板句中的槽
    pattern = re.compile(r"[(].*?[)]")
    slots = re.findall(pattern, template_question)
    # print(slots)
    # 构造SQL语句
    mysql_string = "select * from " + search_table + " where "

    for i_slot in range(len(slots)):
        if i_slot == len(slots) - 1:
            mysql_string += "[" + slots[i_slot][1:-1] + "='" + slots[i_slot] + "'" + "]"
        else:
            mysql_string += "[" + slots[i_slot][1:-1] + "='" + slots[i_slot] + "' and " + "]"
    mysql_string += ";"
    function_logger.info("MySQL语句构造完成！")
    return mysql_string


# 通过模板类型（槽位）构造答句
def build_mysql_answer_string_by_template(template_answer: str, query_result_item: tuple) -> str:
    """
    通过模板类型（槽位）构造mysql答句
    :param template_answer: 模板答句
    :param query_result_item: 本次查询结果（带属性键值对）
    :return:
    """
    # 提取槽位
    pattern = re.compile(r"[(].*?[)]")
    slots = re.findall(pattern, template_answer)
    answer_string = template_answer
    for slot in slots:
        answer_key = query_result_item[slot[1:-1]]
        answer_string = answer_string.replace(slot, str(answer_key))
    return answer_string


# 通过模板类型及关键词键值映射返回mysql语句
# noinspection PyProtectedMember
def build_mysql_string_by_template_and_keymap(template_question: str, template_question_type: str,
                                              keyword_dict: dict) -> str:
    """
    通过模板类型及关键词键值映射返回mysql语句
    :param template_question: 模板问题
    :param template_question_type: 模板问题类型
    :param keyword_dict: 关键词映射
    :return: SQL语句
    """
    function_logger = MyLog(logger=sys._getframe().f_code.co_name).getlog()
    function_logger.info("开始构造MySQL语句...")
    search_table = template_question_type
    # 提取模板句中的槽
    pattern = re.compile(r"[(].*?[)]")
    slots = re.findall(pattern, template_question)
    # print(slots)
    # 构造SQL语句
    mysql_string = ""
    for i_slot in range(len(slots)):
        # function_logger.debug("slot:"+slots[i_slot][1:-1])
        key = keyword_dict["search_" + slots[i_slot][1:-1]]
        # function_logger.debug("key"+key)
        if key == "":
            continue
        else:
            mysql_string += slots[i_slot][1:-1] + "='" + key + "' and "
    if mysql_string != "":
        mysql_string = "select * from " + search_table + " where " + mysql_string[:-5] + ";"
    function_logger.info("MySQL语句构造完成！")
    return mysql_string


# 通过问题模板文件加载问题模板
def load_template_by_file(file_path: str) -> tuple:
    """
    通过问题模板文件加载问题模板
    :param file_path: 模板路径
    :return: 问句条件词、问句目标词、模板答句、模板问句集
    """
    with open(file_path, "rb") as p_file:
        template_dict = pickle.load(p_file)
    return template_dict["fq_conditon"], template_dict["fq_target"], \
           template_dict["ts_answers"], template_dict["ts_questions"]


if __name__ == '__main__':
    main_logger = MyLog(logger=__name__).getlog()
    main_logger.info("start...")
    # build_mysql_string_by_template("(school)(year)(major)在(province)招生人数是多少？")
    # print(build_mysql_string_by_template("(school)(year)(major)在(province)招生人数是多少？"))
    test_template_path = "Template"
    # fq_condition, fq_target, ts_answers, ts_questions \
    #     = load_template_by_file(test_template_path + "/admission_score_major")
    # print(fq_condition)
    # print(fq_target)
    # print(ts_answers)
    # print(ts_questions)
    # test_template_path = "Template"
    build_template_by_fields(test_template_path + "/admission_score_major")
    with open("Template/admission_score_major", "rb")as p_file:
        data = pickle.load(p_file)
    print(data)
    print(len(data["ts_questions"]))
    main_logger.info("end...")
